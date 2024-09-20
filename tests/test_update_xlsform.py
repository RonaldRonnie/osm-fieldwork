# Copyright (c) 2022, 2023 Humanitarian OpenStreetMap Team
#
# This file is part of osm_fieldwork.
#
#     osm-fieldwork is free software: you can redistribute it and/or modify
#     it under the terms of the GNU General Public License as published by
#     the Free Software Foundation, either version 3 of the License, or
#     (at your option) any later version.
#
#     osm-fieldwork is distributed in the hope that it will be useful,
#     but WITHOUT ANY WARRANTY; without even the implied warranty of
#     MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#     GNU General Public License for more details.
#
#     You should have received a copy of the GNU General Public License
#     along with osm_fieldwork.  If not, see <https:#www.gnu.org/licenses/>.
#
"""Test functionality of update_form.py."""

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook, worksheet

from osm_fieldwork.update_xlsform import append_mandatory_fields


async def test_merge_mandatory_fields():
    """Merge the mandatory fields XLSForm to a test survey form."""
    test_form = Path(__file__).parent / "testdata" / "test_form_for_mandatory_fields.xls"

    with open(test_form, "rb") as xlsform:
        form_bytes = BytesIO(xlsform.read())

    updated_form = await append_mandatory_fields(form_bytes, "buildings")
    workbook = load_workbook(filename=BytesIO(updated_form.getvalue()))

    check_survey_sheet(workbook)
    check_choices_sheet(workbook)
    check_entities_sheet(workbook)
    check_form_title(workbook)

    # Write merged xlsform to file for debugging
    with open("merged_xlsform.xlsx", "wb") as merged_xlsform:
        merged_xlsform.write(updated_form.getvalue())


async def test_add_extra_select_from_file():
    """Append extra select_one_from_file questions based on Entity list names."""
    test_form = Path(__file__).parent / "testdata" / "test_form_for_mandatory_fields.xls"

    with open(test_form, "rb") as xlsform:
        form_bytes = BytesIO(xlsform.read())

    updated_form = await append_mandatory_fields(form_bytes, "buildings", additional_entities=["roads", "waterpoints"])
    workbook = load_workbook(filename=BytesIO(updated_form.getvalue()))

    survey_sheet = workbook["survey"]
    name_column = [cell.value for cell in survey_sheet["B"]]

    assert "road" in name_column, "The 'road' field was not added to the survey sheet."
    assert "waterpoint" in name_column, "The 'waterpoint' field was not added to the survey sheet."


async def test_add_task_ids_to_choices():
    """Test appending each task id as a row in choices sheet."""
    test_form = Path(__file__).parent / "testdata" / "test_form_for_mandatory_fields.xls"
    with open(test_form, "rb") as xlsform:
        form_bytes = BytesIO(xlsform.read())

    task_count = 7
    updated_form = await append_mandatory_fields(form_bytes, "buildings", task_count=task_count)
    workbook = load_workbook(filename=BytesIO(updated_form.getvalue()))

    choices_sheet = workbook["choices"]
    # Assuming 'name' is in column B
    name_column = [cell.value for cell in choices_sheet["B"]]

    # Assert each task_id is in the name_column
    task_ids = [1, 2, 3, 4, 5, 6, 7]
    for task_id in task_ids:
        assert task_id in name_column, f"Task ID {task_id} not found in the choices sheet."


def check_survey_sheet(workbook: Workbook) -> None:
    """Check the 'survey' sheet values and ensure no duplicates in 'name' column."""
    survey_sheet = get_sheet(workbook, "survey")
    name_col_index = get_column_index(survey_sheet, "name")
    calculation_col_index = get_column_index(survey_sheet, "calculation")

    form_category_row_index = get_row_index(survey_sheet, name_col_index, "form_category")
    form_category_calculation = survey_sheet.cell(row=form_category_row_index, column=calculation_col_index).value

    expected_calculation = "once('building')"
    assert form_category_calculation == expected_calculation, (
        f"Expected 'calculation' value for 'form_category' to be '{expected_calculation}', "
        f"but got '{form_category_calculation}'."
    )

    check_for_duplicates(survey_sheet, name_col_index)


def check_choices_sheet(workbook: Workbook) -> None:
    """Check the 'choices' sheet and ensure no duplicates in 'name' column."""
    choices_sheet = get_sheet(workbook, "choices")
    name_col_index = get_column_index(choices_sheet, "name")

    check_for_duplicates(choices_sheet, name_col_index)


def check_entities_sheet(workbook: Workbook) -> None:
    """Check the 'entities' sheet values."""
    entities_sheet = get_sheet(workbook, "entities")
    label_col_index = get_column_index(entities_sheet, "label")

    test_label_present = any(
        row[0].value == "test label"
        for row in entities_sheet.iter_rows(min_col=label_col_index, max_col=label_col_index, min_row=2)
    )
    assert not test_label_present, "'test label' found in the 'label' column of 'entities' sheet."


def check_form_title(workbook: Workbook) -> None:
    """Check if the form_title is set correctly in the 'settings' sheet."""
    settings_sheet = get_sheet(workbook, "settings")
    form_title_col_index = get_column_index(settings_sheet, "form_title")

    form_title_value = settings_sheet.cell(row=2, column=form_title_col_index).value
    assert form_title_value == "buildings", "form_title field is not set to 'buildings'"


def get_sheet(workbook: Workbook, sheet_name: str) -> worksheet.worksheet.Worksheet:
    """Helper function to get a sheet or raise an error."""
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"The '{sheet_name}' sheet was not found in the workbook")
    return workbook[sheet_name]


def check_for_duplicates(sheet: worksheet.worksheet.Worksheet, col_index: int) -> None:
    """Check for any duplicate values in a specific column of a sheet."""
    seen_values = set()
    for row in sheet.iter_rows(min_col=col_index, max_col=col_index, min_row=2):
        value = row[0].value
        if value in seen_values:
            raise AssertionError(f"Duplicate value '{value}' found in column '{col_index}' of sheet '{sheet.title}'.")
        seen_values.add(value)


def get_column_index(sheet: worksheet.worksheet.Worksheet, column_name: str) -> int:
    """Get the column index for the given column name."""
    for col_idx, col in enumerate(sheet.iter_cols(1, sheet.max_column), start=1):
        if col[0].value == column_name:
            return col_idx
    raise ValueError(f"Column '{column_name}' not found.")


def get_row_index(sheet: worksheet.worksheet.Worksheet, column_index: int, value: str) -> int:
    """Get the row index where the given column has the specified value."""
    for row_idx, row in enumerate(sheet.iter_rows(min_col=column_index, max_col=column_index), start=1):
        if row[0].value == value:
            return row_idx
    raise ValueError(f"Value '{value}' not found in column {column_index}.")
